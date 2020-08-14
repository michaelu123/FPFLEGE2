import datetime
from decimal import Decimal

import openpyxl
import utils
from kivymd.uix.dialog import MDDialog
from kivymd.uix.list import OneLineListItem
from openpyxl.utils.datetime import CALENDAR_MAC_1904

global conn

# db row names, 0 based
R_Tag = 0
R_Fnr = 1
R_Einsatzstelle = 2
R_Beginn = 3
R_Ende = 4
R_Fahrtzeit = 5
R_Mvv_euro = 6
R_Kh = 7

# Excel column names, 0 based
Tag = 0
Einsatzstelle1 = 1
Beginn1 = 2
Ende1 = 3
Kh1 = 4
Fahrt1 = 5
MVVEuro = 6
Einsatzstelle2 = 7
Beginn2 = 8
Ende2 = 9
Kh2 = 10
Fahrt2 = 11
Einsatzstelle3 = 12
Beginn3 = 13
Ende3 = 14
Kh3 = 15
Arbeitsstunden = 16
Sollstunden = 17
Ueberstunden = 18
Kumuliert = 19
KumFormel = 20
Sentinel = 21

hourFormat = "#,##0.00" # "[hh]:mm" fails on negative times
euroFormat = "#,##0.00€"

class ArbExcel:
    def __init__(self, month, dataDir, app):
        self.month = month
        self.dataDir = dataDir
        self.app = app
        self.tag = None
        self.kumArb = "00:00"
        self.kumSoll = "00:00"
        self.kumUeber = "00:00"
        pass

    def makeExcel(self):
        rows = self.checkComplete()
        if rows is None:
            return
        excelFile = self.writeExcel(rows)
        return excelFile

    def checkRow(self, row):
        beginn = row[R_Beginn]
        ende = row[R_Ende]
        dauer = utils.tsub(ende, beginn)
        if dauer[0] == '-':
            return "Ende vor Beginn"
        if dauer == "00:00":
            return "Beginn gleich Ende"
        if int(dauer[0:2]) >= 12:
            return "Beginn bis Ende mehr als 12 Stunden"
        return None

    def checkComplete(self):
        rows = []
        try:
            with conn:
                c = conn.cursor()
                c.execute("SELECT tag,fnr,einsatzstelle,beginn,ende,fahrtzeit,mvv_euro,kh "
                          "from arbeitsblatt WHERE tag like ?",
                          ("__." + self.month,))
                while True:
                    r = c.fetchmany(100)
                    r = utils.elimEmpty(r, 2)
                    if len(r) == 0:
                        break
                    rows.extend(r)
        except Exception as e:
            utils.printEx("arbex0:", e)

        rows.sort(key=lambda xr: xr[0] + str(xr[1]))  # sort by tag and fnr, i.e. 01.01.20, 02.01.20,...,31.01.20
        lastFnr = 0
        lastTnr = 999
        nrows = []
        for row in rows:
            # print("row", row)
            if row[R_Einsatzstelle] == "" and row[R_Beginn] == "" and row[R_Ende] == "":
                continue
            self.tag = row[R_Tag]
            tnr = utils.tag2Nummer(self.tag)
            wday = utils.day2WT(self.tag)
            if wday == "Sa" or wday == "So":
                continue
            dayMiss = fnrMiss = False
            if lastTnr != 999 and tnr > lastTnr + 1:
                if wday != "Mo":
                    self.tag = utils.num2Tag(lastTnr + 1)
                    wday = utils.day2WT(self.tag)
                    dayMiss = True
            fnr = row[R_Fnr]
            if lastTnr != tnr:
                if fnr != 1:
                    fnrMiss = True
            elif fnr != lastFnr + 1:
                fnrMiss = True
            lastTnr = tnr
            lastFnr = fnr
            if dayMiss or fnrMiss or (row[R_Einsatzstelle] == "" or row[R_Beginn] == "" or row[R_Ende] == ""):
                dia = MDDialog(size_hint=(.8, .4), title="Daten unvollständig",
                               text="Bitte Daten von " + wday + ", " + self.tag + " vervollständigen",
                               text_button_ok="OK", events_callback=self.evcb)
                dia.open()
                return None
            text = self.checkRow(row)
            if text is not None:
                dia = MDDialog(size_hint=(.8, .4), title="Daten falsch",
                               text=text + ", bitte Daten von " + wday + ", " + self.tag + " ausbessern",
                               text_button_ok="OK", events_callback=self.evcb)
                dia.open()
                return None
            nrows.append(row)
        # print("Vollständig!")
        return nrows

    def evcb(self, _x, _y):
        tnr = utils.tag2Nummer(self.tag)
        self.app.gotoScreen(tnr, False)

    def makeRow(self, er, dsum, nsum, uesum, sollstunden, exrownr):
        er[Arbeitsstunden] = utils.hhmm2td(dsum)
        ueberstunden = "00:00"
        if dsum == "00:00" and nsum == "00:00" and uesum == "00:00":
            # nicht möglich
            sollstunden = "00:00"
        elif dsum == "00:00" and nsum == "00:00" and uesum != "00:00":
            ueberstunden = "-" + sollstunden
        elif dsum == "00:00" and nsum != "00:00" and uesum == "00:00":
            sollstunden = "00:00"
        elif dsum == "00:00" and nsum != "00:00" and uesum != "00:00":
            ueberstunden = "-" + uesum
            sollstunden = utils.tsub(sollstunden, nsum)
        elif dsum != "00:00" and nsum == "00:00" and uesum == "00:00":
            ueberstunden = utils.tsub(dsum, sollstunden)
        elif dsum != "00:00" and nsum == "00:00" and uesum != "00:00":
            ueberstunden = utils.tsub(dsum, sollstunden)
        elif dsum != "00:00" and nsum != "00:00" and uesum == "00:00":
            sollstunden = utils.tsub(sollstunden, nsum)
        elif dsum != "00:00" and nsum != "00:00" and uesum != "00:00":
            sollstunden = utils.tsub(sollstunden, nsum)
            ueberstunden = utils.tsub(dsum, sollstunden)

        er[Ueberstunden] = utils.hhmm2td(ueberstunden)
        er[Sollstunden] = utils.hhmm2td(sollstunden)
        self.kumArb = utils.tadd(self.kumArb, dsum)
        self.kumSoll = utils.tadd(self.kumSoll, sollstunden)
        self.kumUeber = utils.tadd(self.kumUeber, ueberstunden)
        er[Kumuliert] = utils.hhmm2td(self.kumUeber)
        er[KumFormel] = "=S" + str(exrownr + 1) + "+U" + str(exrownr)

    def writeExcel(self, rows):
        if self.wb is None:
            wb = openpyxl.Workbook()
            wb.epoch = CALENDAR_MAC_1904  # this enables negative timedeltas
            ws = wb.active
            ws.title = self.month
        else:
            wb = self.wb
            sheetnames = self.wb.get_sheet_names()
            index = sheetnames.index(self.ws.title)
            ws = wb.create_sheet(self.ws.title + "_korr", index + 1)
        ws.append(
            ["Tag", "1.Einsatzstelle", "Beginn", "Ende", "KH", "Fahrt", "MVV",
             "2.Einsatzstelle", "Beginn", "Ende", "KH", "Fahrt",
             "3.Einsatzstelle", "Beginn", "Ende", "KH",
             "Arbeitsstunden", "Sollstunden", "Überstunden", "Kumuliert", "KumFormel"])
        ws.append(
            ["", "Übertrag", "", "", "", "", "",
             "", "", "", "", "",
             "", "", "", "",
             "", "", 0, "=S2", "=S2"])
        exrownr = 2
        ctag = ""
        er = None
        sumh = {}
        sumd = {"Fahrtzeit": 0}
        dsum = "00:00"
        nsum = "00:00"
        fsum = "00:00"
        uesum = "00:00"
        sollStunden = "00:00"
        mvvSumme = Decimal("0.00")
        rows.append(("99",))
        for row in rows:
            tag = row[R_Tag]
            if tag != ctag:
                if ctag != "":
                    self.makeRow(er, dsum, nsum, uesum, sollStunden, exrownr)
                    ws.append(er)
                    exrownr += 1
                    mr = ws.max_row
                    for col in [Arbeitsstunden, Sollstunden, Ueberstunden, Kumuliert, KumFormel]:
                        ws.cell(row=mr, column=col + 1).number_format = hourFormat
                    ws.cell(row=mr, column=MVVEuro + 1).number_format = euroFormat
                if tag == "99":
                    break
                er = ["" for _ in range(Sentinel)]
                wday = utils.day2WT(tag)
                sollStunden = self.app.menu.wtag2Stunden[utils.wday2No[wday]]
                er[Tag] = wday + ", " + tag
                ctag = tag
                dsum = "00:00"
                nsum = "00:00"
                uesum = "00:00"
            fnr = row[R_Fnr]
            es = row[R_Einsatzstelle]
            beginn = row[R_Beginn]
            ende = row[R_Ende]
            kh = row[R_Kh]
            if fnr == 1:
                er[Einsatzstelle1] = es
                er[Beginn1] = beginn
                er[Ende1] = ende
                er[Kh1] = "Ja" if kh else ""
                if row[R_Fahrtzeit] != "":
                    er[Fahrt1] = float(row[R_Fahrtzeit].replace(",", "."))
                    fsum = utils.tadd(fsum, "00:30")
                    dsum = utils.tadd(dsum, "00:30")
                    utils.dadd(sumd, "Fahrtzeit")
                if row[R_Mvv_euro] != "":
                    er[MVVEuro] = float(row[R_Mvv_euro].replace(",", "."))
            elif fnr == 2:
                er[Einsatzstelle2] = es
                er[Beginn2] = beginn
                er[Ende2] = ende
                er[Kh2] = "Ja" if kh else ""
                if row[R_Fahrtzeit] != "":
                    er[Fahrt2] = float(row[R_Fahrtzeit].replace(",", "."))
                    fsum = utils.tadd(fsum, "00:30")
                    dsum = utils.tadd(dsum, "00:30")
                    utils.dadd(sumd, "Fahrtzeit")
            elif fnr == 3:
                er[Einsatzstelle3] = es
                er[Beginn3] = beginn
                er[Ende3] = ende
                er[Kh3] = "Ja" if kh else ""
            else:
                raise ValueError("Tag " + tag + " hat fnr " + fnr)
            utils.hadd(sumh, row)
            utils.dadd(sumd, es)
            dauer = utils.tsubPause(ende, beginn)
            if es.lower() == "üst-abbau":
                uesum = utils.tadd(uesum, dauer)
            elif es.lower() in utils.nichtArb:
                nsum = utils.tadd(nsum, dauer)
            else:
                dsum = utils.tadd(dsum, dauer)
            if row[R_Mvv_euro] != "":
                mvvSumme = mvvSumme + Decimal(row[R_Mvv_euro].replace(",", "."))

        mr = ws.max_row
        mrs = str(mr)
        ws.append([])
        ws.append(["Formeln:", "", "", "", "", "", "=SUM(G2:G" + mrs + ")", "", "", "", "", "", "", "", "", "",
                   "=SUM(Q2:Q" + mrs + ")", "=SUM(R2:R" + mrs + ")", "=SUM(S2:S" + mrs + ")"])
        mr = ws.max_row
        for col in [Arbeitsstunden, Sollstunden, Ueberstunden, Kumuliert]:
            ws.cell(row=mr, column=col + 1).number_format = hourFormat
        ws.cell(row=mr, column=MVVEuro + 1).number_format = euroFormat

        # mvvSumme = utils.moneyfmt(mvvSumme, sep='.', dp=',')
        mvvSumme = utils.moneyfmt(mvvSumme, sep='', dp='.')
        mvvSumme = float(mvvSumme)
        ws.append(["Summen:", "", "", "", "", "", mvvSumme, "", "", "", "", "", "", "", "", "",
                   utils.hhmm2td(self.kumArb), utils.hhmm2td(self.kumSoll), utils.hhmm2td(self.kumUeber)])
        mr = ws.max_row
        for col in [Arbeitsstunden, Sollstunden, Ueberstunden, Kumuliert]:
            ws.cell(row=mr, column=col + 1).number_format = hourFormat
        ws.cell(row=mr, column=MVVEuro + 1).number_format = euroFormat

        ws.append([])
        ws.append(["", "Einsatzstelle", "Tage", "Stunden"])
        sumh["Fahrtzeit"] = fsum
        tsum = "00:00"
        for es in sorted(sumh.keys()):
            if es == "Üst-Abbau":
                continue
            ws.append(("", es, sumd[es], utils.hhmm2td(sumh[es])))
            mr = ws.max_row
            ws.cell(row=mr, column=4).number_format = hourFormat
            if es.lower() not in utils.nichtArb:
                tsum = utils.tadd(tsum, sumh[es])
        ws.append([])
        ws.append(("", "Arbeitsstunden", utils.hhmm2td(tsum)))
        mr = ws.max_row
        ws.cell(row=mr, column=3).number_format = hourFormat
        ws.append(("", "Sollstunden", utils.hhmm2td(self.kumSoll)))
        mr = ws.max_row
        ws.cell(row=mr, column=3).number_format = hourFormat
        ws.append(("", "Überstunden", utils.hhmm2td(self.kumUeber)))
        mr = ws.max_row
        ws.cell(row=mr, column=3).number_format = hourFormat

        # https://bitbucket.org/openpyxl/openpyxl/issues/425/auto_size-not-working
        # for c in range(Sentinel):
        #     dim = ws.column_dimensions[chr(ord("A") + c)]
        #     dim.bestFit = True
        #     dim.customWidth = False
        col_widths = {Tag: 12,
                      Einsatzstelle1: 20, Beginn1: 7, Ende1: 7, Kh1: 3, Fahrt1: 5, MVVEuro: 8,
                      Einsatzstelle2: 20, Beginn2: 7, Ende2: 7, Kh2: 3, Fahrt2: 5,
                      Einsatzstelle3: 20, Beginn3: 7, Ende3: 7, Kh3: 3,
                      Arbeitsstunden: 7, Sollstunden: 7, Ueberstunden: 7, Kumuliert: 7, KumFormel: 7}
        for k in col_widths.keys():
            ws.column_dimensions[chr(ord("A") + k)].width = col_widths[k]

        if self.path is None:
            fn = self.dataDir + "/arbeitsblatt." + self.month + "_" + self.app.menu.ids.vorname.text + "_" + \
                 self.app.menu.ids.nachname.text + ".xlsx"
            wb.save(fn)
            return fn
        else:
            wb.save(self.path)
            return None

    def readExcel(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(filename=path)
        self.wb.epoch = CALENDAR_MAC_1904  # this enables negative timedeltas
        sheetnames = self.wb.get_sheet_names()
        self.dialog = MDDialog(
            title="Wähle Blatt aus",
            type="simple",
            items=[SelectItem(self, t) for t in sheetnames],
        )
        self.dialog.auto_dismiss = False
        self.dialog.open()

    def readExcel2(self, sheetname):
        self.dialog.dismiss()
        print("sheetname", sheetname)
        self.ws = self.wb.get_sheet_by_name(sheetname)
        self.wochenStunden = "00"
        for row in self.ws.rows:
            if len(row) == 0:
                continue
            self.row2db(row)
        self.app.nextScreen(1)
        self.app.nextScreen(-1)

    def row2db(self, row):
        tag = row[Tag].value
        if tag is None or len(tag) != 12:
            return
        tag = tag[4:]  # Mi, 01.07.20 -> 01.07.20
        if len(tag.split(".")) != 3:
            return

        # infer month and first day of month (to become screen "0")
        if self.month == "00.00":
            self.month = tag[3:]  # 07.20
            utils.firstDay = datetime.date(2000 + int(tag[6:8]), int(tag[3:5]), 1)
        elif self.month != tag[3:]:
            return

        # infer wochenstunden from sollstunden
        # see main.py, Menu, setWtagStunden
        if self.wochenStunden == "00":
            sollStunden = self.xval(row[Sollstunden])
            if sollStunden == "04:00":
                self.wochenStunden = "20"
            elif sollStunden == "06:00":
                self.wochenStunden = "30"
            elif sollStunden == "07:00":
                self.wochenStunden = "35"
            elif sollStunden == "08:00" or sollStunden == "06:30":
                self.wochenStunden = "38,5"

            if self.wochenStunden != "00":
                self.app.menu.setWtagStunden(self.wochenStunden)

        if self.xval(row[Einsatzstelle1]) != "":
            try:
                with conn:
                    c = conn.cursor()
                    vals = {"tag": tag, "fnr": 1, "einsatzstelle": row[Einsatzstelle1].value,
                            "beginn": self.xval(row[Beginn1]), "ende": self.xval(row[Ende1]),
                            "fahrtzeit": self.xval(row[Fahrt1]), "mvv_euro": self.xval(row[MVVEuro]),
                            "kh": self.kh(row[Kh1])}
                    c.execute(
                        "INSERT INTO arbeitsblatt VALUES(:tag,:fnr,:einsatzstelle,:beginn,:ende,:fahrtzeit,:mvv_euro,:kh)",
                        vals)
            except Exception as e:
                utils.printEx("fam0:", e)

        if self.xval(row[Einsatzstelle2]) != "":
            try:
                with conn:
                    c = conn.cursor()
                    vals = {"tag": tag, "fnr": 2, "einsatzstelle": row[Einsatzstelle2].value,
                            "beginn": self.xval(row[Beginn2]), "ende": self.xval(row[Ende2]),
                            "fahrtzeit": self.xval(row[Fahrt2]), "mvv_euro": "", "kh": self.kh(row[Kh2])}
                    c.execute(
                        "INSERT INTO arbeitsblatt VALUES(:tag,:fnr,:einsatzstelle,:beginn,:ende,:fahrtzeit,:mvv_euro,:kh)",
                        vals)
            except Exception as e:
                utils.printEx("fam2:", e)

        if self.xval(row[Einsatzstelle3]) != "":
            try:
                with conn:
                    c = conn.cursor()
                    vals = {"tag": tag, "fnr": 3, "einsatzstelle": row[Einsatzstelle3].value,
                            "beginn": self.xval(row[Beginn3]), "ende": self.xval(row[Ende3]),
                            "fahrtzeit": "", "mvv_euro": "", "kh": self.kh(row[Kh1])}
                    c.execute(
                        "INSERT INTO arbeitsblatt VALUES(:tag,:fnr,:einsatzstelle,:beginn,:ende,:fahrtzeit,:mvv_euro,:kh)",
                        vals)
            except Exception as e:
                utils.printEx("fam3:", e)

    def kh(self, c):
        if c.value is None:
            return 0
        if c.value == "Ja":
            return 1
        return 0

    def xval(self, c):
        if c.value is None:
            return ""
        if isinstance(c.value, datetime.time):
            return str(c.value)[0:5]
        return str(c.value)


class SelectItem(OneLineListItem):
    def __init__(self, arbExcel, text):
        super().__init__(text=text)
        self.arbExcel = arbExcel

    def on_press(self, *args):
        self.arbExcel.readExcel2(self.text)
